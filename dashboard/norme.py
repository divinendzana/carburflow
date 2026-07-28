"""Norme de rapport CarburFlow — alignée sur les fiches de suivi terrain."""

from __future__ import annotations

import csv
import io
import unicodedata
from datetime import date, datetime
from typing import Any

from django.db import transaction

from dashboard.models import Rapport

# Colonnes du modèle téléchargeable = fiche de suivi + période.
# Ordre et libellés volontairement proches de data/ligne_rapport.csv.
NORME_COLUMNS = [
    'date_debut',
    'date_fin',
    'id_cuve_principale',
    'id_cuve_journaliere',
    'id_groupe',
    'quantités_cuve_principale',
    'quantite_cuve_journaliere',
    'depotage',
    'compteur_horaire',
    'état_fonctionnement',
    'observations',
]

# Alias acceptés à la lecture (anciens modèles / variantes sans accents).
COLUMN_ALIASES = {
    'quantite_gasoil_cuve_principale': 'quantités_cuve_principale',
    'quantites_cuve_principale': 'quantités_cuve_principale',
    'quantite_cuve_principale': 'quantités_cuve_principale',
    'quantite_gasoil_cuve_journaliere': 'quantite_cuve_journaliere',
    'etat_fonctionnement': 'état_fonctionnement',
    'etat': 'état_fonctionnement',
}


def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize('NFKD', text)
    return ''.join(ch for ch in normalized if not unicodedata.combining(ch))


def canonicalize_header(name: str) -> str:
    raw = str(name or '').strip().lower()
    if not raw:
        return ''
    if raw in COLUMN_ALIASES:
        return COLUMN_ALIASES[raw]
    # Tentative sans accents
    ascii_key = _strip_accents(raw)
    if ascii_key in COLUMN_ALIASES:
        return COLUMN_ALIASES[ascii_key]
    # Match canonique sans accents (ex. etat_fonctionnement ↔ état_fonctionnement)
    for col in NORME_COLUMNS:
        if _strip_accents(col) == ascii_key or col == raw:
            return col
    return raw


def normalize_row_keys(row: dict) -> dict:
    """Renomme les clés d’une ligne vers les noms de colonnes de la fiche."""
    out = {}
    for key, value in (row or {}).items():
        if key is None:
            continue
        canon = canonicalize_header(str(key))
        if not canon:
            continue
        # Première occurrence gagne (évite d’écraser une valeur déjà canonique)
        if canon not in out or (out[canon] in (None, '') and value not in (None, '')):
            out[canon] = value
    return out


NORME_META = {
    'format': 'CarburFlow Fiche de suivi v1',
    'description': (
        'Même tableau que les fiches de suivi terrain. '
        'id_cuve_principale = nom du site (une cuve principale = un site). '
        'id_cuve_journaliere = nom de la cuve journalière. '
        'Ajoutez date_debut / date_fin (identique sur toutes les lignes), puis importez.'
    ),
    'columns': [
        {
            'name': 'date_debut',
            'label': 'Date de début',
            'type': 'date',
            'required': True,
            'example': '13/07/2026',
            'help': 'Début de la période du relevé. Identique sur toutes les lignes.',
        },
        {
            'name': 'date_fin',
            'label': 'Date de fin',
            'type': 'date',
            'required': True,
            'example': '17/07/2026',
            'help': 'Fin de la période du relevé. Identique sur toutes les lignes.',
        },
        {
            'name': 'id_cuve_principale',
            'label': 'Cuve principale (site)',
            'type': 'string',
            'required': False,
            'example': 'BEPANDA INTERNATIONAL',
            'help': 'Nom du site / cuve principale, comme sur la fiche de suivi.',
        },
        {
            'name': 'id_cuve_journaliere',
            'label': 'Cuve journalière',
            'type': 'string',
            'required': False,
            'example': 'BEPANDA INTERNATIONAL',
            'help': 'Nom de la cuve journalière (fiche de suivi).',
        },
        {
            'name': 'id_groupe',
            'label': 'N° groupe électrogène',
            'type': 'int',
            'required': False,
            'example': '1',
            'help': 'Numéro du groupe électrogène (comme sur la fiche).',
        },
        {
            'name': 'quantités_cuve_principale',
            'label': 'Quantité cuve principale (L)',
            'type': 'float',
            'required': False,
            'example': '8448',
            'help': 'Litres mesurés dans la cuve principale.',
        },
        {
            'name': 'quantite_cuve_journaliere',
            'label': 'Quantité cuve journalière (L)',
            'type': 'float',
            'required': False,
            'example': '1000',
            'help': 'Litres mesurés dans la cuve journalière.',
        },
        {
            'name': 'depotage',
            'label': 'Dépotage (L)',
            'type': 'float',
            'required': False,
            'example': '0',
            'help': 'Volume dépoté sur la période (0 si aucun).',
        },
        {
            'name': 'compteur_horaire',
            'label': 'Compteur horaire',
            'type': 'float',
            'required': False,
            'example': '1864',
            'help': 'Relevé du compteur horaire du groupe.',
        },
        {
            'name': 'état_fonctionnement',
            'label': 'État de fonctionnement',
            'type': 'string',
            'required': False,
            'example': 'F',
            'help': 'Code d’état (ex. F = fonctionnement).',
        },
        {
            'name': 'observations',
            'label': 'Observations',
            'type': 'string',
            'required': False,
            'example': 'RAS',
            'help': 'Commentaire libre (incident, maintenance, RAS…).',
        },
    ],
}

SAMPLE_ROWS = [
    {
        'date_debut': '13/07/2026',
        'date_fin': '17/07/2026',
        'id_cuve_principale': 'BEPANDA INTERNATIONAL',
        'id_cuve_journaliere': 'BEPANDA INTERNATIONAL',
        'id_groupe': '1',
        'quantités_cuve_principale': '8448',
        'quantite_cuve_journaliere': '1000',
        'depotage': '0',
        'compteur_horaire': '1864',
        'état_fonctionnement': 'F',
        'observations': 'RAS',
    },
]


def build_csv_bytes(include_sample: bool = True) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=NORME_COLUMNS, lineterminator='\n')
    writer.writeheader()
    if include_sample:
        for row in SAMPLE_ROWS:
            writer.writerow(row)
    return buffer.getvalue().encode('utf-8-sig')


def build_xlsx_bytes(include_sample: bool = True) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rapport'
    ws.append(NORME_COLUMNS)
    if include_sample:
        for row in SAMPLE_ROWS:
            ws.append([row.get(col, '') for col in NORME_COLUMNS])

    meta = wb.create_sheet('Meta')
    meta.append(['champ', 'valeur'])
    meta.append(['format', NORME_META['format']])
    meta.append(['description', NORME_META['description']])
    meta.append(['colonnes', ', '.join(NORME_COLUMNS)])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def rapport_to_rows(rapport: Rapport) -> list[dict]:
    """Reconstitue les lignes au format fiche de suivi pour export."""
    rows = []
    lignes = rapport.lignes.select_related(
        'cuve_principale',
        'cuve_journaliere',
        'groupe_electrogene',
    ).all()
    for ligne in lignes:
        rows.append({
            'date_debut': rapport.date_debut.isoformat() if rapport.date_debut else '',
            'date_fin': rapport.date_fin.isoformat() if rapport.date_fin else '',
            'id_cuve_principale': (
                ligne.cuve_principale.identifiant if ligne.cuve_principale_id else ''
            ),
            'id_cuve_journaliere': (
                ligne.cuve_journaliere.identifiant if ligne.cuve_journaliere_id else ''
            ),
            'id_groupe': str(ligne.groupe_electrogene_id) if ligne.groupe_electrogene_id else '',
            'quantités_cuve_principale': ligne.quantite_gasoil_cuve_principale
            if ligne.quantite_gasoil_cuve_principale is not None
            else '',
            'quantite_cuve_journaliere': ligne.quantite_gasoil_cuve_journaliere
            if ligne.quantite_gasoil_cuve_journaliere is not None
            else '',
            'depotage': ligne.depotage if ligne.depotage is not None else '',
            'compteur_horaire': ligne.compteur_horaire if ligne.compteur_horaire is not None else '',
            'état_fonctionnement': ligne.etat_fonctionnement or '',
            'observations': ligne.observations or '',
        })
    return rows


def build_rapport_csv_bytes(rapport: Rapport) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=NORME_COLUMNS, lineterminator='\n')
    writer.writeheader()
    for row in rapport_to_rows(rapport):
        writer.writerow(row)
    return buffer.getvalue().encode('utf-8-sig')


def build_rapport_xlsx_bytes(rapport: Rapport) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rapport'
    ws.append(NORME_COLUMNS)
    for row in rapport_to_rows(rapport):
        ws.append([row.get(col, '') for col in NORME_COLUMNS])

    meta = wb.create_sheet('Meta')
    meta.append(['champ', 'valeur'])
    meta.append(['rapport_id', rapport.id])
    meta.append(['date_debut', str(rapport.date_debut)])
    meta.append(['date_fin', str(rapport.date_fin)])
    if rapport.created_by_id:
        meta.append(['created_by', rapport.created_by.get_username()])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


COLUMN_LABELS = {
    col['name']: col.get('label') or col['name'] for col in NORME_META['columns']
}


class ImportValidationError(Exception):
    """Erreur d'import avec détails compréhensibles (ligne / colonne)."""

    def __init__(self, message: str, errors: list[dict] | None = None):
        super().__init__(message)
        self.message = message
        self.errors = errors or []

    def as_dict(self) -> dict:
        return {
            'detail': self.message,
            'errors': self.errors,
        }


def _friendly_error(
    *,
    row: int | None,
    column: str | None,
    message: str,
    how_to_fix: str,
) -> dict:
    return {
        'row': row,
        'column': column,
        'column_label': COLUMN_LABELS.get(column or '', column),
        'message': message,
        'how_to_fix': how_to_fix,
    }


def _parse_date(value: Any, *, row: int | None = None, column: str = 'date') -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    label = COLUMN_LABELS.get(column, column)
    raise ImportValidationError(
        f'Une date n’est pas au bon format (ligne {row}).',
        [
            _friendly_error(
                row=row,
                column=column,
                message=f'La valeur « {text} » n’est pas une date reconnue pour « {label} ».',
                how_to_fix='Écrivez la date comme ceci : 13/07/2026 ou 2026-07-13.',
            )
        ],
    )


def _to_float(value: Any, default: float = 0.0, *, row: int | None = None, column: str | None = None) -> float:
    if value is None or value == '':
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(',', '.').replace(' ', '')
    if not text:
        return default
    try:
        return float(text)
    except ValueError as exc:
        label = COLUMN_LABELS.get(column or '', column or 'nombre')
        raise ImportValidationError(
            f'Un nombre est incorrect (ligne {row}).',
            [
                _friendly_error(
                    row=row,
                    column=column,
                    message=f'La valeur « {value} » n’est pas un nombre valide pour « {label} ».',
                    how_to_fix='Mettez uniquement un nombre, par exemple 4500 ou 12,5 (sans lettres).',
                )
            ],
        ) from exc


def _to_int_or_none(value: Any, *, row: int | None = None, column: str | None = None) -> int | None:
    if value is None or value == '':
        return None
    try:
        if isinstance(value, float):
            return int(value)
        text = str(value).strip()
        if not text:
            return None
        return int(float(text))
    except ValueError as exc:
        label = COLUMN_LABELS.get(column or '', column or 'identifiant')
        raise ImportValidationError(
            f'Un numéro est incorrect (ligne {row}).',
            [
                _friendly_error(
                    row=row,
                    column=column,
                    message=f'La valeur « {value} » n’est pas un numéro valide pour « {label} ».',
                    how_to_fix='Mettez uniquement le numéro du groupe (chiffres, sans lettres).',
                )
            ],
        ) from exc


def _to_name_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_headers(headers: list[str]) -> list[str]:
    return [canonicalize_header(h) for h in headers]


def _missing_required_columns(headers: list[str]) -> list[dict]:
    missing = [c for c in ('date_debut', 'date_fin') if c not in headers]
    if not missing:
        return []
    return [
        _friendly_error(
            row=1,
            column=col,
            message=f'La colonne « {COLUMN_LABELS.get(col, col)} » est absente du fichier.',
            how_to_fix='Retéléchargez le modèle (étape 1) et ne renommez pas les titres des colonnes.',
        )
        for col in missing
    ]


def _detect_csv_delimiter(sample: str) -> str:
    """Détecte ; ou , (Excel FR utilise souvent ;)."""
    first_line = (sample.splitlines() or [''])[0]
    if first_line.count(';') >= first_line.count(','):
        return ';'
    return ','


def rows_from_csv(file_bytes: bytes) -> list[dict]:
    try:
        text = file_bytes.decode('utf-8-sig')
    except UnicodeDecodeError as exc:
        raise ImportValidationError(
            'Le fichier CSV n’est pas lisible.',
            [
                _friendly_error(
                    row=None,
                    column=None,
                    message='Le fichier semble mal enregistré (encodage incorrect).',
                    how_to_fix='Enregistrez le fichier en CSV UTF-8, ou utilisez plutôt le modèle Excel.',
                )
            ],
        ) from exc
    delimiter = _detect_csv_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise ImportValidationError(
            'Le fichier CSV est vide ou sans titres de colonnes.',
            [
                _friendly_error(
                    row=1,
                    column=None,
                    message='La première ligne (titres) est manquante.',
                    how_to_fix='Utilisez le modèle téléchargé à l’étape 1 : la 1re ligne doit contenir les noms des colonnes.',
                )
            ],
        )
    headers = _normalize_headers(reader.fieldnames)
    missing_errors = _missing_required_columns(headers)
    if missing_errors:
        raise ImportValidationError(
            'Il manque des colonnes obligatoires dans votre fichier.',
            missing_errors,
        )
    rows = []
    for raw in reader:
        row = normalize_row_keys({k: v for k, v in raw.items() if k})
        if not any(str(v or '').strip() for v in row.values()):
            continue
        rows.append(row)
    return rows


def rows_from_xlsx(file_bytes: bytes) -> list[dict]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ImportValidationError(
            'Le fichier Excel n’a pas pu être ouvert.',
            [
                _friendly_error(
                    row=None,
                    column=None,
                    message='Le fichier est peut-être endommagé ou n’est pas un vrai Excel.',
                    how_to_fix='Retéléchargez le modèle Excel, recopiez vos données dedans, puis réessayez.',
                )
            ],
        ) from exc
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ImportValidationError(
            'Le fichier Excel est vide.',
            [
                _friendly_error(
                    row=1,
                    column=None,
                    message='Aucune ligne trouvée dans la première feuille.',
                    how_to_fix='Utilisez le modèle de l’étape 1 et ajoutez vos lignes sous les titres.',
                )
            ],
        ) from exc
    headers = _normalize_headers(list(header_row))
    missing_errors = _missing_required_columns(headers)
    if missing_errors:
        raise ImportValidationError(
            'Il manque des colonnes obligatoires dans votre fichier.',
            missing_errors,
        )
    rows = []
    for values in rows_iter:
        if values is None or all(v is None or str(v).strip() == '' for v in values):
            continue
        raw = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            raw[key] = values[idx] if idx < len(values) else None
        rows.append(normalize_row_keys(raw))
    return rows


@transaction.atomic
def import_report_rows(rows: list[dict], user, *, create_missing: bool = False) -> tuple[Rapport, int]:
    """
    Import web / API : délègue au pipeline (analyse + insertion).
    Par défaut refuse les IDs inconnus (create_missing=False).
    """
    from dashboard.rapport_pipeline import import_rapport_lignes

    normalized = [normalize_row_keys(r) for r in rows]
    result = import_rapport_lignes(
        normalized,
        user=user,
        create_missing=create_missing,
        require_entities=True,
    )
    rapport = Rapport.objects.get(pk=result.rapport_id)
    return rapport, result.imported_lines
